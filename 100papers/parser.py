#!/usr/bin/env python3
# -*- coding: utf-8 -*-
#
# Géry Casiez - 2023
#

import bibtexparser
from openpyxl import Workbook

with open('100papers/biblio.bib') as bibtex_file:
    bibtex_str = bibtex_file.read()

bib = bibtexparser.loads(bibtex_str)
bibEntries = bib.entries

p = 0

wb = Workbook()
ws = wb.active

for b in bibEntries:
    if p < 100 and (b['ENTRYTYPE'] == 'article' or b['ENTRYTYPE'] == 'inproceedings'):
        w = ""
        if 'journal' in b.keys():
            w = b['journal']
        elif 'booktitle' in b.keys():
            w = b['booktitle']
        ws['A%s'%(p+1)].value = b['ID']
        ws['B%s'%(p+1)].value = b['ENTRYTYPE']
        ws['C%s'%(p+1)].value = b['author']
        ws['D%s'%(p+1)].value = b['title']
        ws['D%s'%(p+1)].hyperlink = "https://scholar.google.com/scholar?q=%s"%(b['title'].replace(" ", "+"))
        ws['E%s'%(p+1)].value = w
        print("%s %s %s %s %s %s"%(p, b['ENTRYTYPE'], b['ID'], b['author'], b['title'], w))
        p += 1

wb.save('papers.xlsx')